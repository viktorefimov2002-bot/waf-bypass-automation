from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .classifier import load_groups
from .common import read_jsonl


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="DDEBF7")


def _excel_value(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _append(ws, values: list[Any]) -> None:
    ws.append([_excel_value(value) for value in values])


def _style_sheet(ws, widths: dict[int, int] | None = None) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 30
    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for cell in ws[row]:
                cell.fill = ALT_FILL
        for cell in ws[row]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, ws.max_column + 1):
        width = widths.get(column) if widths else None
        if width is None:
            width = min(50, max(10, max(len(str(ws.cell(row, column).value or "")) for row in range(1, min(ws.max_row, 200) + 1)) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width
    if ws.max_row >= 2 and ws.max_column >= 1:
        table = Table(displayName=f"Table_{ws.title.replace(' ', '_').replace('-', '_')}", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        ws.add_table(table)


def create_report(input_path: Path, groups_path: Path, output_path: Path, taxonomy_path: Path | None = None) -> dict[str, Any]:
    records = read_jsonl(input_path)
    groups = load_groups(groups_path, taxonomy_path)
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Summary")
    _append(ws, ["Metric", "Value"])
    _append(ws, ["Variants", len(records)])
    _append(ws, ["Payload files", len({r['payload_path'] for r in records})])
    _append(ws, ["BYPASS by code", sum(r.get("http_code") != 403 for r in records)])
    _append(ws, ["Confirmed origin", sum(r.get("final_verdict") == "BYPASS_ORIGIN_CONFIRMED" for r in records)])
    _append(ws, ["Groups represented", len({r.get('group_id') for r in records if r.get('group_id') is not None})])
    _style_sheet(ws, {1: 30, 2: 24})

    payloads: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        payloads[record["payload_path"]].append(record)
    ws = wb.create_sheet("Payload mapping")
    _append(ws, ["Payload", "Category", "Variants", "Zones", "Encodings", "Group ID", "Group", "Type", "Confidence", "Reason"])
    for path in sorted(payloads):
        rows = payloads[path]
        first = rows[0]
        _append(ws, [
            path,
            first["category"],
            len(rows),
            ", ".join(sorted({r["zone"] for r in rows})),
            ", ".join(sorted({r["encoding"] for r in rows})),
            first.get("group_id"),
            first.get("group_name"),
            first.get("classification_type"),
            first.get("classification_confidence"),
            first.get("classification_reason"),
        ])
    _style_sheet(ws, {1: 24, 2: 12, 3: 10, 4: 24, 5: 20, 6: 10, 7: 42, 8: 18, 9: 14, 10: 60})

    ws = wb.create_sheet("Bypass variants")
    headers = [
        "Payload", "Category", "Variant", "Zone", "Encoding", "HTTP code", "Code verdict", "Server", "Route verdict", "Final verdict",
        "Group ID", "Group", "Normalized payload", "cURL hash", "cURL",
    ]
    _append(ws, headers)
    for record in records:
        _append(ws, [
            record["payload_path"], record["category"], record["variant"], record["zone"], record["encoding"], record.get("http_code"),
            record.get("code_verdict"), record.get("server_header"), record.get("route_verdict"), record.get("final_verdict"), record.get("group_id"),
            record.get("group_name"), record.get("normalized_payload"), record.get("curl_hash"), record.get("curl"),
        ])
    _style_sheet(ws, {1: 22, 2: 12, 3: 20, 4: 13, 5: 13, 6: 11, 7: 20, 8: 20, 9: 20, 10: 30, 11: 10, 12: 40, 13: 60, 14: 20, 15: 90})

    counts = Counter((r.get("group_id"), r.get("group_name")) for r in records)
    payload_counts: dict[tuple[Any, Any], set[str]] = defaultdict(set)
    for record in records:
        payload_counts[(record.get("group_id"), record.get("group_name"))].add(record["payload_path"])
    ws = wb.create_sheet("Group summary")
    _append(ws, ["Group ID", "Group", "Payload files", "Variants", "Source"])
    represented = set()
    for key, count in sorted(counts.items(), key=lambda item: (-item[1], str(item[0]))):
        group_id, name = key
        represented.add(group_id)
        source = groups.get(group_id, {}).get("source", "unclassified") if group_id is not None else "unclassified"
        _append(ws, [group_id, name, len(payload_counts[key]), count, source])
    for group_id in sorted(groups):
        if group_id not in represented:
            _append(ws, [group_id, groups[group_id]["name"], 0, 0, groups[group_id]["source"]])
    _style_sheet(ws, {1: 10, 2: 60, 3: 14, 4: 12, 5: 18})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {"output": str(output_path), "sheets": wb.sheetnames, "variants": len(records)}
